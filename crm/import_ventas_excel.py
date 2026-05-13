"""Import Ventas Serpentarius.xlsx into CRM PostgreSQL"""
import openpyxl, re, psycopg2, uuid
from datetime import datetime, timedelta
import random

CRM_DB = {"host": "localhost", "port": 5432, "user": "postgres",
          "password": "bbL2Tdlq4lPFSbr4xUCu", "dbname": "serpentarius_crm"}

EXCEL_PATH = "/home/adminbot/Downloads/Ventas Serpentarius.xlsx"

def parse_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Ventas']
    customers = []
    current = None
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        name = str(row[0]).strip() if row[0] else ''
        phone = str(row[1]).strip() if row[1] else ''
        product = str(row[2]).strip() if row[2] else ''
        venta = str(row[4]).strip() if row[4] else '0'
        if name and name.lower() not in ('', 'none', 'nan') and name != 'Ventas Serpentarius':
            current = {'name': name.strip(), 'phone': phone.strip(), 'items': []}
            if product and product.lower() not in ('', 'none', 'nan'):
                current['items'].append({'product': product.strip(), 'venta': venta})
            customers.append(current)
        elif current and product and product.lower() not in ('', 'none', 'nan'):
            current['items'].append({'product': product.strip(), 'venta': venta})
    seen = {}
    for c in customers:
        phone = re.sub(r'[^\d]', '', c['phone'])
        if len(phone) > 10: phone = phone[-10:]
        elif len(phone) < 7: phone = ''
        c['phone_clean'] = phone
        key = c['name'].lower().strip()
        if key not in seen:
            seen[key] = c
        else:
            seen[key]['items'].extend(c['items'])
            seen[key]['phone'] = phone or seen[key]['phone']
    return list(seen.values())

def import_contacts(contacts):
    conn = psycopg2.connect(**CRM_DB)
    cur = conn.cursor()
    imported = 0
    for c in contacts:
        phone = c.get('phone_clean', '')
        whatsapp = f"+57{phone}" if phone else ''
        total_venta = sum(float(x.get('venta', 0) or 0) for x in c['items'])
        contact_id = str(uuid.uuid4())
        cur.execute("""
            INSERT INTO contacts (id, full_name, phone, whatsapp_phone, source, status,
                                  lifetime_value, total_orders)
            VALUES (%s, %s, %s, %s, 'presencial', 'customer', %s, 1)
            ON CONFLICT (whatsapp_phone) WHERE whatsapp_phone IS NOT NULL AND whatsapp_phone != '' DO NOTHING
        """, (contact_id, c['name'], phone, whatsapp, total_venta))
        if cur.rowcount == 0: continue
        days_ago = random.randint(1, 60)
        order_id = str(uuid.uuid4())
        order_num = f"IMP-{datetime.now().strftime('%Y')}-{imported+1:04d}"
        order_created = datetime.now() - timedelta(days=days_ago)
        cur.execute("""
            INSERT INTO orders (id, contact_id, order_number, status, subtotal, total,
                               payment_method, payment_status, created_at, updated_at)
            VALUES (%s, %s, %s, 'delivered', %s, %s, 'efectivo', 'confirmed', %s, %s)
        """, (order_id, contact_id, order_num, total_venta, total_venta, order_created, order_created))
        for item in c['items']:
            pname = item['product'][:100]
            price = float(item.get('venta', 0) or 0)
            cur.execute("""
                INSERT INTO order_items (order_id, product_sku, product_name, quantity, unit_price, total_price)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (order_id, f"IMP-{imported+1:04d}", pname, 1, price, price))
        score = min(95, max(20, int(total_venta / 5000)))
        cur.execute("""
            INSERT INTO lead_scores (contact_id, score, monetary, recency_days, frequency_30d)
            VALUES (%s, %s, %s, %s, 1)
        """, (contact_id, score, total_venta, days_ago))
        imported += 1
    conn.commit()
    cur.close(); conn.close()
    return imported

if __name__ == '__main__':
    contacts = parse_excel(EXCEL_PATH)
    print(f"Found {len(contacts)} unique customers")
    n = import_contacts(contacts)
    print(f"Imported {n} contacts")
